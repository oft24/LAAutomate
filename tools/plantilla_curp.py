"""Crea el Excel de entrada de la automatización `curp_desde_excel`.

    python tools/plantilla_curp.py

Escribe `datos/personas.xlsx` con las columnas exactas que la consulta
necesita y dos filas de ejemplo. No sobrescribe un archivo existente: si ya
tienes tus datos ahí, perderlos por correr esto dos veces sería un precio
absurdo por una plantilla.

Las columnas no son opcionales ni intercambiables. El formulario de
gob.mx/curp pide los apellidos POR SEPARADO y la entidad de nacimiento, que
no se puede deducir del nombre ni de la fecha.
"""
from __future__ import annotations

import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

DESTINO = RAIZ / "datos" / "personas.xlsx"

COLUMNAS = [
    ("nombres", "Nombre(s) de pila, como aparecen en el acta"),
    ("primer_apellido", "Apellido paterno"),
    ("segundo_apellido", "Apellido materno (déjalo vacío si no tiene)"),
    ("dia", "Día de nacimiento, 1-31"),
    ("mes", "Mes de nacimiento, 1-12"),
    ("anio", "Año de nacimiento, 4 dígitos"),
    ("sexo", "Hombre / Mujer / No binario  (o H / M / X)"),
    ("estado", "Estado de nacimiento: nombre o clave de 2 letras (JC, DF, NE...)"),
]

EJEMPLOS = [
    {
        "nombres": "Juan Carlos",
        "primer_apellido": "Pérez",
        "segundo_apellido": "Ramírez",
        "dia": 5,
        "mes": 3,
        "anio": 1990,
        "sexo": "Hombre",
        "estado": "Jalisco",
    },
    {
        "nombres": "María Fernanda",
        "primer_apellido": "López",
        "segundo_apellido": "",
        "dia": 22,
        "mes": 11,
        "anio": 1985,
        "sexo": "Mujer",
        "estado": "DF",
    },
]


def main() -> int:
    if DESTINO.exists():
        print(f"Ya existe {DESTINO} — no se toca.")
        print("Bórralo a mano si de verdad quieres regenerar la plantilla.")
        return 1

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    DESTINO.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "personas"

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="1F4E5F")
    for columna, (nombre, ayuda) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=columna, value=nombre)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")
        # La explicación va como comentario de la celda, no en una fila
        # aparte: una fila de ayuda se leería como una persona más.
        celda.comment = None
        hoja.column_dimensions[get_column_letter(columna)].width = max(len(nombre) + 4, 16)

    for fila, ejemplo in enumerate(EJEMPLOS, start=2):
        for columna, (nombre, _ayuda) in enumerate(COLUMNAS, start=1):
            hoja.cell(row=fila, column=columna, value=ejemplo[nombre])

    hoja.freeze_panes = "A2"

    # Segunda hoja con el significado de cada columna, para que la
    # plantilla se explique sola sin ensuciar los datos.
    guia = libro.create_sheet("guía")
    guia["A1"], guia["B1"] = "columna", "qué va aquí"
    guia["A1"].font = guia["B1"].font = Font(bold=True)
    for fila, (nombre, ayuda) in enumerate(COLUMNAS, start=2):
        guia.cell(row=fila, column=1, value=nombre)
        guia.cell(row=fila, column=2, value=ayuda)
    guia.column_dimensions["A"].width = 20
    guia.column_dimensions["B"].width = 70

    libro.save(DESTINO)
    print(f"Plantilla creada: {DESTINO}")
    print("Borra las dos filas de ejemplo y pon tus datos.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
