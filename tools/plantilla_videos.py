"""Crea el Excel de búsquedas de `buscar_videos_youtube`.

    python tools/plantilla_videos.py

Escribe `datos/videos_buscar.xlsx` con dos columnas y unas filas de
ejemplo. No sobrescribe un archivo existente: si ya tienes tus búsquedas
ahí, perderlas por correr esto dos veces sería un precio absurdo.

Para añadir búsquedas, escribe filas nuevas y vuelve a ejecutar la
automatización: las que ya tienen resultados se saltan solas.
"""
from __future__ import annotations

import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

DESTINO = RAIZ / "datos" / "videos_buscar.xlsx"

COLUMNAS = [
    ("tema", "Qué buscar. Obligatorio."),
    ("canal", "Canal concreto, opcional. Se añade al texto de búsqueda."),
]

EJEMPLOS = [
    {"tema": "automatización con python", "canal": ""},
    {"tema": "selenium tutorial", "canal": "Código Espinoza"},
    {"tema": "power automate vs python", "canal": ""},
]


def main() -> int:
    if DESTINO.exists():
        print(f"Ya existe {DESTINO} — no se toca.")
        print("Añade filas ahí y vuelve a ejecutar la automatización:")
        print("solo se buscarán las nuevas.")
        return 1

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    DESTINO.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "busquedas"

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="1F4E5F")
    for columna, (nombre, _ayuda) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=columna, value=nombre)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")
        hoja.column_dimensions[get_column_letter(columna)].width = 38

    for fila, ejemplo in enumerate(EJEMPLOS, start=2):
        for columna, (nombre, _ayuda) in enumerate(COLUMNAS, start=1):
            hoja.cell(row=fila, column=columna, value=ejemplo[nombre])

    hoja.freeze_panes = "A2"

    guia = libro.create_sheet("guía")
    guia["A1"], guia["B1"] = "columna", "qué va aquí"
    guia["A1"].font = guia["B1"].font = Font(bold=True)
    for fila, (nombre, ayuda) in enumerate(COLUMNAS, start=2):
        guia.cell(row=fila, column=1, value=nombre)
        guia.cell(row=fila, column=2, value=ayuda)
    guia.cell(row=len(COLUMNAS) + 3, column=1, value="cómo añadir")
    guia.cell(
        row=len(COLUMNAS) + 3,
        column=2,
        value="Escribe filas nuevas y vuelve a ejecutar la automatización. "
        "Las búsquedas que ya tienen resultados se saltan; solo se consultan las nuevas.",
    )
    guia.column_dimensions["A"].width = 20
    guia.column_dimensions["B"].width = 80

    libro.save(DESTINO)
    print(f"Plantilla creada: {DESTINO}")
    print("Cambia los ejemplos por tus búsquedas y ejecuta «buscar_videos_youtube».")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
