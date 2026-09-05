"""Busca vídeos en YouTube a partir de un Excel de temas y canales.

Lee `datos/videos_buscar.xlsx` (una fila por búsqueda: tema + canal
opcional), consulta YouTube y escribe `datos/videos_encontrados.xlsx` con
los resultados de cada una.

**Añade filas cuando quieras y vuelve a ejecutarla**: las búsquedas que ya
tienen resultados se saltan, así que solo se consultan las nuevas. Para
rehacer una, borra sus filas del Excel de salida (o el archivo entero).

Sobre YouTube: se leen los resultados públicos de búsqueda como lo haría
una persona, con una pausa entre consultas. No hay sesión ni datos de
cuenta. Si necesitas volumen de verdad, la API de datos de YouTube es el
camino correcto — esto es automatización de navegador y depende de que la
página no cambie.
"""
from __future__ import annotations

import time
from pathlib import Path
from urllib.parse import quote_plus

from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By

from core.config import BASE_DIR, var
from engine.automation_base import AutomationResult, BaseAutomation
from engine.registry import registrar

EXCEL_ENTRADA = BASE_DIR / "datos" / "videos_buscar.xlsx"
EXCEL_SALIDA = BASE_DIR / "datos" / "videos_encontrados.xlsx"

# Cuántos vídeos se guardan por búsqueda y cuánto se espera entre ellas.
MAX_POR_BUSQUEDA = int(var("YT_MAX_RESULTADOS", "5"))
PAUSA_ENTRE_BUSQUEDAS = float(var("YT_PAUSA", "3"))

COLUMNAS_MINIMAS = ("tema",)

# Lo que lleva la plantilla que se crea si el Excel no existe.
COLUMNAS_PLANTILLA = [
    ("tema", "Qué buscar. Obligatorio."),
    ("canal", "Canal concreto, opcional. Se añade al texto de búsqueda."),
]
EJEMPLOS_PLANTILLA = [
    {"tema": "automatización con python", "canal": ""},
    {"tema": "selenium tutorial", "canal": "Código Espinoza"},
    {"tema": "power automate vs python", "canal": ""},
]

# Selectores verificados contra la página real el 2026-09-04. YouTube
# convive con componentes viejos (ytd-*) y nuevos (yt-*-view-model); estos
# son los que devuelven datos hoy.
RESULTADO = "ytd-video-renderer"
TITULO = "a#video-title"
CANAL = "ytd-channel-name a"
METADATOS = "#metadata-line span"
DURACION = "#text.ytd-thumbnail-overlay-time-status-renderer, badge-shape div"

# Cuánto se espera a que aparezca el primer resultado. NO es un sleep fijo:
# se espera a que el elemento exista. Consultar el DOM antes de que YouTube
# termine de renderizar devuelve cero resultados y parece "no encontré
# nada" cuando en realidad la página aún no había pintado.
ESPERA_RESULTADOS = 15


def normalizar(valor) -> str:
    """Texto limpio de una celda, tratando el vacio de Excel como vacio.

    Una celda vacia llega desde pandas como `float("nan")`, y `str(nan)`
    es la cadena "nan". Sin este filtro, esa palabra viaja como dato: la
    busqueda de YouTube acabo consultando literalmente "automatizacion con
    python nan", y en la del CURP se habria mandado "nan" como segundo
    apellido a un servicio oficial.
    """
    if valor is None:
        return ""
    # NaN es el unico valor que no es igual a si mismo.
    if isinstance(valor, float) and valor != valor:
        return ""
    texto = str(valor).strip()
    return "" if texto.lower() == "nan" else texto


def construir_consulta(fila: dict) -> str:
    """El texto que se busca en YouTube para esta fila.

    El canal se añade como texto libre y no con el filtro de canal de
    YouTube: el filtro exige el identificador exacto, que nadie tiene a
    mano al escribir un Excel, mientras que el nombre suelto funciona con
    lo que la gente sí sabe escribir.
    """
    tema = normalizar(fila.get("tema"))
    canal = normalizar(fila.get("canal"))
    if not tema and not canal:
        raise ValueError("la fila no tiene ni tema ni canal")
    return f"{tema} {canal}".strip()


def clave_de_busqueda(fila: dict) -> str:
    """Identifica una búsqueda para poder saltarla si ya se hizo.

    En minúsculas y sin espacios de más: «Python  » y «python» son la misma
    búsqueda, y repetirla solo gastaría tiempo.
    """
    return " ".join(construir_consulta(fila).lower().split())


def url_de_busqueda(consulta: str) -> str:
    return f"https://www.youtube.com/results?search_query={quote_plus(consulta)}"


def crear_plantilla(destino: Path) -> None:
    """Escribe el Excel de búsquedas con su cabecera y unos ejemplos.

    Vive aquí y no solo en `tools/plantilla_videos.py` porque la app
    instalada no trae Python ni esa carpeta: decirle a alguien que corra un
    script que no tiene es un callejón sin salida.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    destino.parent.mkdir(parents=True, exist_ok=True)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "busquedas"
    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="1F4E5F")
    for columna, (nombre, _ayuda) in enumerate(COLUMNAS_PLANTILLA, start=1):
        celda = hoja.cell(row=1, column=columna, value=nombre)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")
        hoja.column_dimensions[get_column_letter(columna)].width = 38

    for fila, ejemplo in enumerate(EJEMPLOS_PLANTILLA, start=2):
        for columna, (nombre, _ayuda) in enumerate(COLUMNAS_PLANTILLA, start=1):
            hoja.cell(row=fila, column=columna, value=ejemplo[nombre])
    hoja.freeze_panes = "A2"

    guia = libro.create_sheet("guía")
    guia["A1"], guia["B1"] = "columna", "qué va aquí"
    guia["A1"].font = guia["B1"].font = Font(bold=True)
    for fila, (nombre, ayuda) in enumerate(COLUMNAS_PLANTILLA, start=2):
        guia.cell(row=fila, column=1, value=nombre)
        guia.cell(row=fila, column=2, value=ayuda)
    guia.cell(row=len(COLUMNAS_PLANTILLA) + 3, column=1, value="cómo añadir")
    guia.cell(
        row=len(COLUMNAS_PLANTILLA) + 3,
        column=2,
        value="Escribe filas nuevas y vuelve a ejecutar la automatización. "
        "Las búsquedas que ya tienen resultados se saltan; solo se consultan las nuevas.",
    )
    guia.column_dimensions["A"].width = 20
    guia.column_dimensions["B"].width = 80
    libro.save(destino)


@registrar(nombre="buscar_videos_youtube", disparador="manual", categoria="investigacion")
class BuscarVideosYoutube(BaseAutomation):
    def ejecutar(self) -> AutomationResult:
        if not EXCEL_ENTRADA.exists():
            crear_plantilla(EXCEL_ENTRADA)
            return AutomationResult(
                success=True,
                message=(
                    f"No había búsquedas todavía, así que creé la plantilla en "
                    f"{EXCEL_ENTRADA}. Ábrela, escribe tus temas en la columna «tema» "
                    "y vuelve a ejecutar."
                ),
                data={"plantilla_creada": str(EXCEL_ENTRADA)},
            )

        filas = self.excel.leer(EXCEL_ENTRADA)
        if not filas:
            return AutomationResult(success=True, message="El Excel no tiene búsquedas.")

        faltantes = [c for c in COLUMNAS_MINIMAS if c not in filas[0]]
        if faltantes:
            raise ValueError(
                f"Al Excel le falta la columna: {', '.join(faltantes)}. "
                "Se esperan «tema» y, opcionalmente, «canal»."
            )

        ya_hechas = self._busquedas_previas()
        resultados = self._resultados_previos()
        nuevas = hechas = vacias = 0

        for numero, fila in enumerate(filas, start=2):
            try:
                consulta = construir_consulta(fila)
            except ValueError as exc:
                self.logger.warning("Fila %s: %s", numero, exc)
                continue

            if clave_de_busqueda(fila) in ya_hechas:
                self.logger.info("Fila %s (%s): ya buscada, se salta", numero, consulta)
                continue

            nuevas += 1
            self.logger.info("Fila %s: buscando «%s»", numero, consulta)
            try:
                encontrados = self._buscar(consulta)
            except Exception as exc:  # noqa: BLE001 - una búsqueda mala no tumba el lote
                self.logger.exception("Fila %s falló", numero)
                resultados.append(
                    {
                        "consulta": consulta,
                        "tema": normalizar(fila.get("tema")),
                        "canal_buscado": normalizar(fila.get("canal")),
                        "posicion": 0,
                        "titulo": "",
                        "canal": "",
                        "url": "",
                        "vistas": "",
                        "publicado": "",
                        "duracion": "",
                        "estado": f"{type(exc).__name__}: {exc}",
                    }
                )
                continue

            if not encontrados:
                vacias += 1
                self.logger.info("Sin resultados para «%s»", consulta)
                resultados.append(
                    {
                        "consulta": consulta,
                        "tema": normalizar(fila.get("tema")),
                        "canal_buscado": normalizar(fila.get("canal")),
                        "posicion": 0,
                        "titulo": "",
                        "canal": "",
                        "url": "",
                        "vistas": "",
                        "publicado": "",
                        "duracion": "",
                        "estado": "sin resultados",
                    }
                )
            else:
                hechas += 1
                for video in encontrados:
                    resultados.append(
                        {
                            "consulta": consulta,
                            "tema": normalizar(fila.get("tema")),
                            "canal_buscado": normalizar(fila.get("canal")),
                            **video,
                            "estado": "ok",
                        }
                    )

            # Se guarda tras CADA búsqueda: si el lote se corta a la mitad,
            # lo ya encontrado no se pierde y la próxima vez se salta.
            self._guardar(resultados)
            time.sleep(PAUSA_ENTRE_BUSQUEDAS)

        if nuevas == 0:
            return AutomationResult(
                success=True,
                message="No había búsquedas nuevas: todas las filas ya tenían resultados.",
                data={"salida": str(EXCEL_SALIDA)},
            )

        self._guardar(resultados)
        return AutomationResult(
            success=True,
            message=f"{hechas} búsqueda(s) con resultados, {vacias} sin ninguno. "
            f"{len(resultados)} filas en total.",
            data={"nuevas": nuevas, "con_resultados": hechas, "vacias": vacias,
                  "salida": str(EXCEL_SALIDA)},
        )

    # ------------------------------------------------------------ pasos

    def _buscar(self, consulta: str) -> list[dict]:
        self.web.ir_a(url_de_busqueda(consulta))
        self._aceptar_o_rechazar_cookies()

        try:
            # Esperar al ELEMENTO, no un tiempo fijo: si se consulta el DOM
            # antes de que YouTube pinte, salen cero resultados y parece
            # "no encontré nada".
            self.web.leer_texto(TITULO, by=By.CSS_SELECTOR)
        except (TimeoutException, WebDriverException):
            return []

        videos = []
        for posicion, elemento in enumerate(
            self.web.driver.find_elements(By.CSS_SELECTOR, RESULTADO)[:MAX_POR_BUSQUEDA], start=1
        ):
            datos = self._leer_resultado(elemento, posicion)
            if datos:
                videos.append(datos)
        return videos

    def _leer_resultado(self, elemento, posicion: int) -> dict | None:
        """Extrae un vídeo. Devuelve None si le falta lo esencial.

        Cada campo va en su propio try: que YouTube deje de mostrar las
        vistas no es razón para perder el título y la URL, que es lo que
        de verdad se viene a buscar.
        """

        def texto(selector: str) -> str:
            """Lee un campo, prefiriendo textContent sobre .text.

            `.text` de Selenium solo devuelve el texto que considera
            RENDERIZADO. El nombre del canal de YouTube salia vacio en
            todas las filas aunque estaba en el DOM: `.text` daba "" y
            `textContent` el valor correcto. Se comprobo con cinco
            selectores distintos, todos con el mismo resultado.
            """
            try:
                hallado = elemento.find_element(By.CSS_SELECTOR, selector)
            except (WebDriverException, Exception):  # noqa: BLE001
                return ""
            try:
                contenido = (hallado.get_attribute("textContent") or "").strip()
                return contenido or hallado.text.strip()
            except (WebDriverException, Exception):  # noqa: BLE001
                return ""

        try:
            enlace = elemento.find_element(By.CSS_SELECTOR, TITULO)
            titulo = (
                enlace.get_attribute("title")
                or enlace.get_attribute("textContent")
                or enlace.text
                or ""
            ).strip()
            url = (enlace.get_attribute("href") or "").split("&")[0]
        except (WebDriverException, Exception):  # noqa: BLE001
            return None
        if not titulo or not url:
            return None

        metadatos = []
        try:
            metadatos = [
                (m.get_attribute("textContent") or m.text or "").strip()
                for m in elemento.find_elements(By.CSS_SELECTOR, METADATOS)
            ]
            metadatos = [m for m in metadatos if m]
        except WebDriverException:
            # Los metadatos (vistas, antiguedad) son opcionales: si YouTube
            # cambia ese bloque, la fila sale sin ellos en vez de perderse.
            pass

        return {
            "posicion": posicion,
            "titulo": titulo,
            "canal": texto(CANAL),
            "url": url,
            "vistas": metadatos[0] if metadatos else "",
            "publicado": metadatos[1] if len(metadatos) > 1 else "",
            "duracion": texto(DURACION),
        }

    def _aceptar_o_rechazar_cookies(self) -> None:
        """Si aparece el aviso de cookies, se elige RECHAZAR.

        Se busca por texto y no por un id concreto: el diálogo cambia de
        maquetación por región y por idioma. Que no aparezca es lo normal
        fuera de la UE, así que no encontrarlo no es un fallo.
        """
        for etiqueta in ("Rechazar todo", "Reject all", "Rechazar"):
            try:
                boton = self.web.driver.find_element(
                    By.XPATH, f"//button[.//*[contains(text(), '{etiqueta}')] or contains(., '{etiqueta}')]"
                )
                if boton.is_displayed():
                    boton.click()
                    self.logger.info("Aviso de cookies rechazado")
                    time.sleep(1)
                    return
            except (WebDriverException, Exception):  # noqa: BLE001
                continue

    # ---------------------------------------------------------- persistencia

    def _resultados_previos(self) -> list[dict]:
        if not EXCEL_SALIDA.exists():
            return []
        try:
            return self.excel.leer(EXCEL_SALIDA)
        except Exception:  # noqa: BLE001 - un archivo corrupto no debe impedir buscar
            self.logger.warning("No pude leer %s; se empieza de cero", EXCEL_SALIDA)
            return []

    def _busquedas_previas(self) -> set[str]:
        """Las consultas que ya tienen alguna fila en la salida.

        Es lo que hace que añadir filas al Excel y volver a ejecutar solo
        consulte las nuevas.
        """
        previas = set()
        for fila in self._resultados_previos():
            consulta = normalizar(fila.get("consulta"))
            if consulta:
                previas.add(" ".join(consulta.lower().split()))
        return previas

    def _guardar(self, resultados: list[dict]) -> None:
        if resultados:
            EXCEL_SALIDA.parent.mkdir(parents=True, exist_ok=True)
            self.excel.escribir(EXCEL_SALIDA, resultados, hoja="videos")
