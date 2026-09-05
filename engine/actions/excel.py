"""Lectura/escritura de Excel. pandas para datos, pywin32 cuando se necesita
controlar Excel abierto (macros, formato condicional en vivo, etc)."""
from __future__ import annotations

from pathlib import Path

import pandas as pd


class ExcelActions:
    def __init__(self, logger) -> None:
        self.logger = logger

    def leer(self, ruta: str | Path, hoja: str | int = 0) -> list[dict]:
        df = pd.read_excel(ruta, sheet_name=hoja)
        return df.to_dict(orient="records")

    def escribir(self, ruta: str | Path, filas: list[dict], hoja: str = "Sheet1") -> None:
        pd.DataFrame(filas).to_excel(ruta, sheet_name=hoja, index=False)
        self.logger.info("Excel escrito: %s (%d filas)", ruta, len(filas))

    def com(self):
        """Acceso a Excel via COM (pywin32) para casos que requieren la app abierta."""
        import win32com.client  # import perezoso: solo Windows con Office instalado

        return win32com.client.Dispatch("Excel.Application")

    def formatear_comparativo(self, ruta: str | Path) -> None:
        """Presenta un reporte de compras ya escrito; conserva sus valores."""
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.formatting.rule import FormulaRule

        libro = load_workbook(ruta)
        try:
            hoja = libro['Comparativo']
            hoja.sheet_view.showGridLines = False
            hoja.freeze_panes = 'D2'
            hoja.auto_filter.ref = hoja.dimensions
            anchos = {'A': 15, 'B': 31, 'C': 18, 'D': 12, 'E': 21, 'F': 23,
                      'G': 19, 'H': 12, 'I': 24, 'J': 44, 'K': 24}
            for columna, ancho in anchos.items():
                hoja.column_dimensions[columna].width = ancho
            for celda in hoja[1]:
                celda.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                celda.fill = PatternFill('solid', fgColor='15283C')
                celda.alignment = Alignment(wrap_text=True, vertical='center')
            hoja.row_dimensions[1].height = 32
            for fila in hoja.iter_rows(min_row=2):
                hoja.row_dimensions[fila[0].row].height = 32
                for celda in fila:
                    celda.font = Font(name='Arial', size=11, color='192B3B')
                    celda.alignment = Alignment(vertical='center', wrap_text=True)
                    if celda.column in (5, 6, 7):
                        celda.number_format = '"$"#,##0.00'
            hoja.conditional_formatting.add(f'A2:I{hoja.max_row}', FormulaRule(
                formula=['$I2="MEJOR OPCIÓN"'], fill=PatternFill('solid', fgColor='DDF5E8'),
                font=Font(bold=True, color='126640')))
            libro.save(ruta)
        finally:
            libro.close()
